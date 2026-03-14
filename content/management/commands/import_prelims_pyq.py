"""
Management command to import Prelims PYQs from Excel files.

Supports TWO Excel formats:

FORMAT A — "Flat" (single sheet, question_id in column):
    Sheet: "Questions" (or any single sheet)
    Columns: question_id, chapter_code, stem, option_a, option_b, option_c,
             option_d, correct_answer, explanation, difficulty, exam_source,
             year, question_type, tags

FORMAT B — "Multi-tab" (one tab per chapter):
    Tabs: "MCQs - Ch1 CDoI", "MCQs - Ch2 APBC", etc.
    Columns: Q.No., Question, Option (a)-(d), Answer, Exam Name,
             Exam Stage, Year, Parse_Error, Review_Status

Auto-detects format from headers in row 1.

Usage:
    python manage.py import_prelims_pyq file.xlsx --dry-run --subject Polity
    python manage.py import_prelims_pyq file.xlsx --batch-id GC --subject Polity
"""

import os
import re
from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from content.models import PrelimsPYQ, QuestionAppearance, Subject


# Answer cleaning map
ANSWER_MAP = {
    '(a)': 'A', '(b)': 'B', '(c)': 'C', '(d)': 'D', '(*)': '*',
    'a': 'A', 'b': 'B', 'c': 'C', 'd': 'D',
}


def clean_answer(raw):
    """(a) -> A, (b) -> B, etc."""
    if not raw:
        return ''
    val = str(raw).strip().lower()
    return ANSWER_MAP.get(val, str(raw).strip().upper())


def clean_year(raw):
    """Convert to int, default 0."""
    try:
        return int(float(raw)) if raw else 0
    except (ValueError, TypeError):
        return 0


def compute_option_avg(opts):
    """Average length of non-empty options."""
    non_empty = [o for o in opts if o]
    if not non_empty:
        return 0
    return sum(len(o) for o in non_empty) // len(non_empty)


class Command(BaseCommand):
    help = 'Import Prelims PYQs from an Excel file into the database'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Path to the Excel file')
        parser.add_argument('--dry-run', action='store_true',
                            help='Preview import without writing to database')
        parser.add_argument('--batch-id', type=str, default='',
                            help='Batch ID prefix (e.g. GC). If empty, derived from filename.')
        parser.add_argument('--subject', type=str, default='',
                            help='Subject name to link (e.g. Polity). Optional.')

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        dry_run = options['dry_run']
        batch_id = options['batch_id']
        subject_name = options['subject']

        # Validate file
        if not os.path.exists(excel_path):
            self.stderr.write(self.style.ERROR(f'File not found: {excel_path}'))
            return

        filename = Path(excel_path).stem

        # Derive batch_id from filename if not given
        if not batch_id:
            match = re.search(r'PY_(\w+)_', filename)
            if match:
                batch_id = match.group(1)
            else:
                batch_id = filename[:10]

        # Lookup subject (exact match first, then partial)
        subject_obj = None
        if subject_name:
            try:
                subject_obj = Subject.objects.get(name__iexact=subject_name)
            except Subject.DoesNotExist:
                # Try partial match: "Polity" matches "Indian Polity & Governance"
                matches = Subject.objects.filter(name__icontains=subject_name)
                if matches.count() == 1:
                    subject_obj = matches.first()
                elif matches.count() > 1:
                    self.stdout.write(self.style.WARNING(
                        f'  WARNING: Multiple subjects match "{subject_name}": '
                        f'{", ".join(m.name for m in matches)}. Using first.'
                    ))
                    subject_obj = matches.first()
            if subject_obj:
                self.stdout.write(
                    f'  Subject: {subject_obj.name} (id={subject_obj.pk})')
            else:
                self.stdout.write(self.style.WARNING(
                    f'  WARNING: Subject "{subject_name}" not found. '
                    f'Continuing with subject=None.'
                ))

        # Load workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        mode_label = 'DRY RUN' if dry_run else 'IMPORT'
        self.stdout.write(f'\n{"=" * 50}')
        self.stdout.write(f'  {mode_label}: {Path(excel_path).name}')
        self.stdout.write(f'  Batch: {batch_id}')
        self.stdout.write(f'{"=" * 50}\n')

        # Detect format: check if any sheet starts with "MCQs -" (multi-tab)
        # or if first data sheet has question_id/stem columns (flat)
        has_mcq_tabs = any(s.startswith('MCQs -') for s in wb.sheetnames)

        if has_mcq_tabs:
            fmt = 'multitab'
            mcq_count = sum(1 for s in wb.sheetnames if s.startswith('MCQs -'))
            self.stdout.write(f'  Format: MULTI-TAB ({mcq_count} MCQ tabs)\n')
        else:
            # Check first sheet headers for flat format
            first_sheet = wb[wb.sheetnames[0]]
            header_row = next(first_sheet.iter_rows(
                min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip().lower() if h else '' for h in header_row]

            if 'question_id' in headers and 'stem' in headers:
                fmt = 'flat'
                self.stdout.write('  Format: FLAT (question_id + stem columns)\n')
            elif 'q.no.' in headers or 'question' in headers:
                fmt = 'multitab'
                self.stdout.write('  Format: MULTI-TAB (MCQs tabs)\n')
            else:
                self.stderr.write(self.style.ERROR(
                    f'  Unknown format. Headers: {headers[:8]}'
                ))
                wb.close()
                return

        # Track totals
        totals = {
            'imported': 0, 'updated': 0, 'skipped': 0,
            'errors': 0, 'appearances': 0,
        }
        skip_reasons = {}

        if fmt == 'flat':
            self._import_flat(wb, batch_id, subject_obj, dry_run,
                              totals, skip_reasons)
        else:
            self._import_multitab(wb, batch_id, subject_obj, dry_run,
                                  totals, skip_reasons)

        wb.close()

        # Final summary
        action = 'would import' if dry_run else 'imported'
        self.stdout.write(f'\n{"=" * 50}')
        self.stdout.write(
            f'  TOTAL: {totals["imported"]:,} {action}'
            f' | {totals["updated"]:,} updated'
            f' | {totals["skipped"]:,} skipped'
            f' | {totals["errors"]:,} errors'
            f' | {totals["appearances"]:,} appearances'
        )
        self.stdout.write(f'{"=" * 50}')

        if skip_reasons:
            self.stdout.write('\nSkipped reasons:')
            for reason, count in sorted(skip_reasons.items()):
                self.stdout.write(f'  {reason}: {count}')

        if dry_run:
            self.stdout.write(self.style.WARNING(
                '\n  DRY RUN — no database writes made.'
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'\n  Import complete. Batch: {batch_id}'
            ))

    # ── FORMAT A: Flat single-sheet ──────────────────────

    def _import_flat(self, wb, batch_id, subject_obj, dry_run,
                     totals, skip_reasons):
        """
        Flat format: single sheet with question_id already in column.
        Columns: question_id, chapter_code, stem, option_a-d,
                 correct_answer, explanation, difficulty, exam_source,
                 year, question_type, tags
        """
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Read header to build column map
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            col_map = {}
            for idx, h in enumerate(header_row):
                if h:
                    col_map[str(h).strip().lower()] = idx

            tab_imported = 0
            tab_updated = 0
            tab_skipped = 0
            tab_errors = 0
            tab_appearances = 0

            rows = list(ws.iter_rows(min_row=2, values_only=True))

            for row_idx, row in enumerate(rows, start=2):
                try:
                    def cell(col_name):
                        idx = col_map.get(col_name)
                        if idx is not None and idx < len(row) and row[idx] is not None:
                            return str(row[idx]).strip()
                        return ''

                    question_id = cell('question_id')
                    stem = cell('stem')
                    opt_a = cell('option_a')
                    opt_b = cell('option_b')
                    opt_c = cell('option_c')
                    opt_d = cell('option_d')
                    correct_answer = clean_answer(cell('correct_answer'))
                    explanation = cell('explanation')
                    difficulty = cell('difficulty') or 'Unknown'
                    exam_source = cell('exam_source') or 'UPPCS'
                    year = clean_year(cell('year'))
                    tags = cell('tags')

                    # Skip empty
                    if not stem.strip():
                        tab_skipped += 1
                        skip_reasons['empty_question'] = skip_reasons.get(
                            'empty_question', 0) + 1
                        continue

                    if not question_id:
                        tab_skipped += 1
                        skip_reasons['no_question_id'] = skip_reasons.get(
                            'no_question_id', 0) + 1
                        continue

                    # Auto-compute
                    stem_length = len(stem)
                    option_avg_length = compute_option_avg(
                        [opt_a, opt_b, opt_c, opt_d])

                    fields = {
                        'stem': stem,
                        'option_a': opt_a,
                        'option_b': opt_b,
                        'option_c': opt_c,
                        'option_d': opt_d,
                        'correct_answer': correct_answer,
                        'explanation': explanation,
                        'difficulty': difficulty,
                        'exam_source': exam_source,
                        'year': year,
                        'tags': tags,
                        'review_status': 'ok',
                        'batch_id': batch_id,
                        'is_free': True,
                        'is_active': True,
                        'stem_length': stem_length,
                        'option_avg_length': option_avg_length,
                        'subject': subject_obj,
                    }

                    if dry_run:
                        tab_imported += 1
                    else:
                        obj, created = PrelimsPYQ.objects.update_or_create(
                            question_id=question_id,
                            defaults=fields,
                        )
                        if created:
                            tab_imported += 1
                        else:
                            tab_updated += 1

                        # QuestionAppearance
                        if year > 0 and exam_source:
                            _, qa_created = QuestionAppearance.objects.get_or_create(
                                question=obj,
                                year=year,
                                exam_source=exam_source,
                                defaults={
                                    'exam_name': exam_source,
                                    'is_primary': True,
                                },
                            )
                            if qa_created:
                                tab_appearances += 1

                except Exception as e:
                    tab_errors += 1
                    self.stderr.write(self.style.ERROR(
                        f'  ERROR row {row_idx} in {sheet_name}: {e}'
                    ))

            action = 'would import' if dry_run else 'imported'
            update_str = f', {tab_updated} updated' if tab_updated else ''
            appear_str = (f', {tab_appearances} appearances'
                          if tab_appearances else '')
            self.stdout.write(
                f'  {sheet_name}: {tab_imported} {action}{update_str}'
                f', {tab_skipped} skipped, {tab_errors} errors{appear_str}'
            )

            totals['imported'] += tab_imported
            totals['updated'] += tab_updated
            totals['skipped'] += tab_skipped
            totals['errors'] += tab_errors
            totals['appearances'] += tab_appearances

    # ── FORMAT B: Multi-tab MCQs ─────────────────────────

    def _import_multitab(self, wb, batch_id, subject_obj, dry_run,
                         totals, skip_reasons):
        """
        Multi-tab format: "MCQs - Ch1 CDoI" tabs.
        Columns: Q.No., Question, Option (a)-(d), Answer,
                 Exam Name, Exam Stage, Year, Parse_Error, Review_Status
        """
        for sheet_name in wb.sheetnames:
            # Skip non-MCQ tabs
            if not sheet_name.startswith('MCQs -'):
                if sheet_name not in (
                    'Progress Tracker', 'Parse Issues', 'Sheet1', 'Sheet'
                ):
                    self.stdout.write(f'  SKIP tab: {sheet_name}')
                continue

            # Extract chapter code: "MCQs - Ch1 CDoI" -> "CDoI"
            code_match = re.search(r'MCQs\s*-\s*Ch\d+\s+(.+)', sheet_name)
            chapter_code = (code_match.group(1).strip() if code_match
                            else sheet_name.replace('MCQs - ', '').strip())

            ws = wb[sheet_name]
            tab_imported = 0
            tab_updated = 0
            tab_skipped = 0
            tab_errors = 0
            tab_appearances = 0

            rows = list(ws.iter_rows(min_row=2, values_only=True))

            for row_idx, row in enumerate(rows, start=2):
                try:
                    def cell(idx):
                        if idx < len(row) and row[idx] is not None:
                            return str(row[idx]).strip()
                        return ''

                    qno_raw = cell(0)
                    question = cell(1)
                    opt_a = cell(2)
                    opt_b = cell(3)
                    opt_c = cell(4)
                    opt_d = cell(5)
                    answer_raw = cell(6)
                    exam_name = cell(7)
                    exam_stage = cell(8)
                    year_raw = cell(9)
                    review_status = cell(11) if len(row) > 11 else ''

                    # Skip
                    if not question.strip():
                        tab_skipped += 1
                        skip_reasons['empty_question'] = skip_reasons.get(
                            'empty_question', 0) + 1
                        continue

                    if review_status == 'blank_question':
                        tab_skipped += 1
                        skip_reasons['blank_question'] = skip_reasons.get(
                            'blank_question', 0) + 1
                        continue

                    # Generate question_id
                    try:
                        qno = int(float(qno_raw)) if qno_raw else row_idx
                    except (ValueError, TypeError):
                        qno = row_idx
                    question_id = f'P-PY-{batch_id}-{chapter_code}-{qno:03d}'

                    correct_answer = clean_answer(answer_raw)
                    year = clean_year(year_raw)
                    stem_length = len(question)
                    option_avg_length = compute_option_avg(
                        [opt_a, opt_b, opt_c, opt_d])

                    if not review_status or review_status not in (
                        'ok', 'needs_review', 'parse_error_no_options',
                        'blank_question', 'draft'
                    ):
                        review_status = 'ok'

                    fields = {
                        'stem': question,
                        'option_a': opt_a,
                        'option_b': opt_b,
                        'option_c': opt_c,
                        'option_d': opt_d,
                        'correct_answer': correct_answer,
                        'exam_source': exam_name if exam_name else 'UPPCS',
                        'year': year,
                        'difficulty': 'Unknown',
                        'review_status': review_status,
                        'batch_id': f'PY-{batch_id}',
                        'is_free': True,
                        'is_active': True,
                        'stem_length': stem_length,
                        'option_avg_length': option_avg_length,
                        'subject': subject_obj,
                    }

                    if dry_run:
                        tab_imported += 1
                    else:
                        obj, created = PrelimsPYQ.objects.update_or_create(
                            question_id=question_id,
                            defaults=fields,
                        )
                        if created:
                            tab_imported += 1
                        else:
                            tab_updated += 1

                        # QuestionAppearance
                        if year > 0 and exam_name:
                            full_exam_name = f'{exam_name} {exam_stage}'.strip()
                            _, qa_created = QuestionAppearance.objects.get_or_create(
                                question=obj,
                                year=year,
                                exam_source=exam_name,
                                defaults={
                                    'exam_name': full_exam_name,
                                    'is_primary': True,
                                },
                            )
                            if qa_created:
                                tab_appearances += 1

                except Exception as e:
                    tab_errors += 1
                    self.stderr.write(self.style.ERROR(
                        f'  ERROR row {row_idx} in {sheet_name}: {e}'
                    ))

            action = 'would import' if dry_run else 'imported'
            update_str = f', {tab_updated} updated' if tab_updated else ''
            appear_str = (f', {tab_appearances} appearances'
                          if tab_appearances else '')
            self.stdout.write(
                f'  {sheet_name}: {tab_imported} {action}{update_str}'
                f', {tab_skipped} skipped, {tab_errors} errors{appear_str}'
            )

            totals['imported'] += tab_imported
            totals['updated'] += tab_updated
            totals['skipped'] += tab_skipped
            totals['errors'] += tab_errors
            totals['appearances'] += tab_appearances
