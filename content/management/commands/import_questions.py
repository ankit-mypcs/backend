from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from content.models import Question, Chapter
import openpyxl
from pathlib import Path
import json


class Command(BaseCommand):
    help = 'Import questions from Excel file (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Validate only, do not save to database'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='Specific sheet name to import (default: first sheet)'
        )

    def handle(self, *args, **options):
        file_path = Path(options['file_path'])
        dry_run = options['dry_run']
        sheet_name = options['sheet']

        # Color codes
        GREEN = '\033[92m'
        RED = '\033[91m'
        YELLOW = '\033[93m'
        BLUE = '\033[94m'
        RESET = '\033[0m'

        # Validate file exists
        if not file_path.exists():
            raise CommandError(f'{RED}File not found: {file_path}{RESET}')

        if file_path.suffix != '.xlsx':
            raise CommandError(f'{RED}File must be .xlsx format{RESET}')

        self.stdout.write(f'{BLUE}Loading Excel file: {file_path}{RESET}')

        # Load workbook
        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            raise CommandError(f'{RED}Error loading Excel file: {e}{RESET}')

        # Select sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(
                    f'{RED}Sheet "{sheet_name}" not found. '
                    f'Available sheets: {", ".join(workbook.sheetnames)}{RESET}'
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        self.stdout.write(f'{BLUE}Using sheet: {sheet.title}{RESET}')

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        # Expected columns
        expected_columns = [
            'question_id', 'chapter_code', 'stem',
            'option_a', 'option_b', 'option_c', 'option_d',
            'correct_answer', 'explanation',
            'difficulty', 'exam_source', 'year', 'question_type', 'tags'
        ]

        # Validate headers
        missing_columns = set(expected_columns) - set(headers)
        if missing_columns:
            raise CommandError(
                f'{RED}Missing required columns: {", ".join(missing_columns)}{RESET}'
            )

        # Statistics
        imported_count = 0
        skipped_count = 0
        errors = []

        # Cache chapters for lookup
        chapters_cache = {ch.code: ch for ch in Chapter.objects.all()}

        self.stdout.write(f'{BLUE}Starting validation and import...{RESET}\n')

        # Process rows (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Create dict from row
            row_data = dict(zip(headers, row))

            # Skip empty rows
            if not any(row_data.values()):
                continue

            # Run quality checks
            row_errors = self._validate_row(row_data, row_num, chapters_cache)

            if row_errors:
                skipped_count += 1
                errors.extend(row_errors)
                continue

            # If not dry run, create the question
            if not dry_run:
                try:
                    self._create_question(row_data, chapters_cache)
                    imported_count += 1
                except Exception as e:
                    skipped_count += 1
                    errors.append(f'Row {row_num}: Database error - {str(e)}')
            else:
                imported_count += 1

        # Print colored report
        self.stdout.write('\n' + '=' * 70)
        self.stdout.write(f'{BLUE}IMPORT SUMMARY{RESET}')
        self.stdout.write('=' * 70)

        if dry_run:
            self.stdout.write(f'{YELLOW}DRY RUN MODE - No data was saved{RESET}')

        self.stdout.write(f'{GREEN}✓ Valid rows: {imported_count}{RESET}')
        self.stdout.write(f'{RED}✗ Skipped rows: {skipped_count}{RESET}')

        if errors:
            self.stdout.write(f'\n{RED}ERRORS:{RESET}')
            for error in errors:
                self.stdout.write(f'  {RED}• {error}{RESET}')

        self.stdout.write('=' * 70)

        if not dry_run and imported_count > 0:
            self.stdout.write(
                f'\n{GREEN}Successfully imported {imported_count} question(s){RESET}'
            )

    def _validate_row(self, row_data, row_num, chapters_cache):
        """Run 10 quality checks on a row"""
        errors = []

        # Check 1: Required fields not empty
        required_fields = [
            'question_id', 'chapter_code', 'stem', 'option_a', 'option_b',
            'option_c', 'option_d', 'correct_answer', 'explanation',
            'difficulty', 'exam_source', 'year', 'question_type'
        ]
        for field in required_fields:
            if not row_data.get(field):
                errors.append(f'Row {row_num}: Missing required field "{field}"')

        # Check 2: Duplicate question_id
        if row_data.get('question_id'):
            if Question.objects.filter(question_id=row_data['question_id']).exists():
                errors.append(
                    f'Row {row_num}: Duplicate question_id "{row_data["question_id"]}"'
                )

        # Check 3: Valid chapter_code exists in DB
        chapter_code = row_data.get('chapter_code')
        if chapter_code and chapter_code not in chapters_cache:
            errors.append(
                f'Row {row_num}: Chapter code "{chapter_code}" not found in database'
            )

        # Check 4: correct_answer is A/B/C/D
        correct_answer = str(row_data.get('correct_answer', '')).upper()
        if correct_answer not in ['A', 'B', 'C', 'D']:
            errors.append(
                f'Row {row_num}: correct_answer must be A, B, C, or D (got "{correct_answer}")'
            )

        # Check 5: Valid difficulty
        difficulty = str(row_data.get('difficulty', '')).lower()
        if difficulty not in ['easy', 'medium', 'hard']:
            errors.append(
                f'Row {row_num}: difficulty must be easy, medium, or hard (got "{difficulty}")'
            )

        # Check 6: Year between 1990-2026
        try:
            year = int(row_data.get('year', 0))
            if year < 1990 or year > 2026:
                errors.append(
                    f'Row {row_num}: year must be between 1990 and 2026 (got {year})'
                )
        except (ValueError, TypeError):
            errors.append(f'Row {row_num}: year must be a valid integer')

        # Check 7: All 4 options non-empty
        for opt in ['option_a', 'option_b', 'option_c', 'option_d']:
            if not row_data.get(opt):
                errors.append(f'Row {row_num}: {opt} cannot be empty')

        # Check 8: exam_source not empty
        if not row_data.get('exam_source'):
            errors.append(f'Row {row_num}: exam_source cannot be empty')

        # Check 9: explanation min 20 chars
        explanation = row_data.get('explanation', '')
        if explanation and len(str(explanation)) < 20:
            errors.append(
                f'Row {row_num}: explanation must be at least 20 characters'
            )

        # Check 10: Valid question_type
        question_type = str(row_data.get('question_type', '')).lower()
        if question_type not in ['factual', 'conceptual', 'application']:
            errors.append(
                f'Row {row_num}: question_type must be factual, conceptual, '
                f'or application (got "{question_type}")'
            )

        return errors

    def _create_question(self, row_data, chapters_cache):
        """Create a Question object from validated row data"""
        # Get chapter
        chapter = chapters_cache[row_data['chapter_code']]

        # Process tags (comma-separated string → list)
        tags_str = row_data.get('tags', '')
        if tags_str:
            tags = [tag.strip() for tag in str(tags_str).split(',') if tag.strip()]
        else:
            tags = []

        # Create question
        Question.objects.create(
            chapter=chapter,
            question_id=row_data['question_id'],
            stem=row_data['stem'],
            option_a=row_data['option_a'],
            option_b=row_data['option_b'],
            option_c=row_data['option_c'],
            option_d=row_data['option_d'],
            correct_answer=str(row_data['correct_answer']).upper(),
            explanation=row_data['explanation'],
            difficulty=str(row_data['difficulty']).lower(),
            exam_source=row_data['exam_source'],
            year=int(row_data['year']),
            question_type=str(row_data['question_type']).lower(),
            tags=tags,
        )
