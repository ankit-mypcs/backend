"""
content/management/commands/upload_chapter.py

Reads a chapter XLSX (10 sheets) and creates all content records.
Header-based column detection — handles Ch4/Ch5/Ch6 inconsistencies.

Usage:
  python manage.py upload_chapter path/to/HIS_StoneAge_Ch4.xlsx \
      --chapter-slug stone-age \
      --chapter-name "The Stone Age" \
      --chapter-number 4 \
      --unit-slug prehistoric-india \
      --dry-run

  python manage.py upload_chapter path/to/HIS_StoneAge_Ch4.xlsx \
      --chapter-slug stone-age \
      --chapter-name "The Stone Age" \
      --chapter-number 4 \
      --unit-slug prehistoric-india
"""
import re
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
import openpyxl

from content.models import (
    Chapter, Unit, Topic, SubTopic,
    Fact, Site, TimelineEvent, GlossaryTerm,
    ExamIntelEntry, ComparisonMatrix, Visual, Exercise, State,
)


def clean(val):
    """Normalise a cell value → stripped string or ''."""
    if val is None:
        return ''
    return str(val).strip()


def header_map(ws):
    """Return {normalised_header: col_index} from row 1."""
    mapping = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
        if cell.value:
            key = clean(cell.value).lower()
            mapping[key] = idx
    return mapping


def find_col(hmap, *candidates):
    """Return index for the first matching candidate header (case-insensitive)."""
    for c in candidates:
        key = c.lower().strip()
        if key in hmap:
            return hmap[key]
    return None


def rows_as_dicts(ws):
    """Yield each data row as {header_key: value} using header_map."""
    hmap = header_map(ws)
    headers_by_idx = {v: k for k, v in hmap.items()}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        d = {}
        for idx, val in enumerate(row):
            if idx in headers_by_idx:
                d[headers_by_idx[idx]] = clean(val)
        if all(v == '' for v in d.values()):
            continue  # skip empty rows
        yield d


class Command(BaseCommand):
    help = 'Upload a chapter XLSX (10 sheets) into the database.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to chapter XLSX file')
        parser.add_argument('--chapter-slug', required=True, help='Chapter slug')
        parser.add_argument('--chapter-name', required=True, help='Chapter display name')
        parser.add_argument('--chapter-number', required=True, type=int, help='Chapter number')
        parser.add_argument('--unit-slug', required=True, help='Slug of parent Unit (must exist)')
        parser.add_argument('--dry-run', action='store_true', help='Preview without writing to DB')

    def handle(self, *args, **options):
        xlsx_path = options['xlsx_path']
        dry_run = options['dry_run']

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Cannot open {xlsx_path}: {e}")

        expected_sheets = [
            '1_Timeline', '2_Concepts', '3_Sites', '4_KeyFacts',
            '5_State_Specific', '6_Society', '7_Images', '8_Terms',
            '9_ExamAnalysis', '10_Exercises',
        ]
        missing = [s for s in expected_sheets if s not in wb.sheetnames]
        if missing:
            raise CommandError(f"Missing sheets: {missing}")

        # Resolve Unit
        try:
            unit = Unit.objects.get(slug=options['unit_slug'])
        except Unit.DoesNotExist:
            raise CommandError(
                f"Unit with slug '{options['unit_slug']}' not found. "
                f"Create taxonomy first (Subject → Part → Unit).")

        chapter_slug = options['chapter_slug']
        chapter_name = options['chapter_name']
        chapter_number = options['chapter_number']

        counters = {
            'timeline': 0, 'concepts': 0, 'sites': 0,
            'keyfacts': 0, 'state_specific': 0, 'society': 0,
            'images': 0, 'terms': 0, 'exam_analysis': 0,
            'exercises': 0, 'topics_created': 0, 'subtopics_created': 0,
        }

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — nothing will be saved."))
            self._dry_run(wb, expected_sheets, chapter_name)
            wb.close()
            return

        with transaction.atomic():
            # ── Create / get Chapter ─────────────────
            chapter, ch_created = Chapter.objects.get_or_create(
                slug=chapter_slug,
                defaults={
                    'unit': unit,
                    'name': chapter_name,
                    'chapter_number': chapter_number,
                    'sort_order': chapter_number,
                },
            )
            if not ch_created:
                self.stdout.write(self.style.WARNING(
                    f"Chapter '{chapter_slug}' already exists (pk={chapter.pk}). "
                    f"Appending content."))

            # Topic / SubTopic helper (get-or-create, cached)
            _topic_cache = {}
            _subtopic_cache = {}

            def get_topic(name_raw):
                name = clean(name_raw)
                if not name:
                    return None
                if name in _topic_cache:
                    return _topic_cache[name]
                obj, created = Topic.objects.get_or_create(
                    chapter=chapter, name=name,
                    defaults={'sort_order': len(_topic_cache)},
                )
                _topic_cache[name] = obj
                if created:
                    counters['topics_created'] += 1
                return obj

            def get_subtopic(topic_obj, name_raw):
                name = clean(name_raw)
                if not name or not topic_obj:
                    return None
                key = (topic_obj.pk, name)
                if key in _subtopic_cache:
                    return _subtopic_cache[key]
                obj, created = SubTopic.objects.get_or_create(
                    topic=topic_obj, name=name,
                    defaults={'sort_order': len(_subtopic_cache)},
                )
                _subtopic_cache[key] = obj
                if created:
                    counters['subtopics_created'] += 1
                return obj

            # ── Sheet 1: Timeline ────────────────────
            ws = wb['1_Timeline']
            for order, d in enumerate(rows_as_dicts(ws)):
                topic_obj = get_topic(d.get('topic', ''))
                TimelineEvent.objects.create(
                    chapter=chapter,
                    date_text=d.get('period/date', ''),
                    event=d.get('event/development', ''),
                    citation=d.get('citation', ''),
                    topic=topic_obj,
                    sort_order=order,
                )
                counters['timeline'] += 1

            # ── Sheet 2: Concepts (matrix) ───────────
            ws = wb['2_Concepts']
            hmap = header_map(ws)
            # Concept columns = everything except 'parameter', 'topic', 'citation'
            skip_keys = {'parameter', 'topic', 'citation'}
            concept_cols = [k for k in hmap if k not in skip_keys]
            concept_cols.sort(key=lambda k: hmap[k])  # sort by column position

            # Group rows by topic → one ComparisonMatrix per topic
            from collections import defaultdict
            matrix_groups = defaultdict(lambda: {'parameters': [], 'data': {}})
            for d in rows_as_dicts(ws):
                topic_name = d.get('topic', 'General')
                param = d.get('parameter', '')
                if not param:
                    continue
                grp = matrix_groups[topic_name]
                grp['parameters'].append(param)
                for col_key in concept_cols:
                    if col_key not in grp['data']:
                        grp['data'][col_key] = {}
                    grp['data'][col_key][param] = d.get(col_key, '')

            for topic_name, grp in matrix_groups.items():
                topic_obj = get_topic(topic_name)
                # Find the citation from any row (usually last column)
                citation = ''
                for d in rows_as_dicts(ws):
                    if d.get('topic', '') == topic_name and d.get('citation', ''):
                        citation = d.get('citation', '')
                        break

                # Store column display names (proper case from header)
                col_display = []
                for ck in concept_cols:
                    # Find original header from worksheet
                    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                        if cell.value and clean(cell.value).lower() == ck:
                            col_display.append(clean(cell.value))
                            break

                ComparisonMatrix.objects.create(
                    chapter=chapter,
                    title=topic_name,
                    parameters=grp['parameters'],
                    columns=col_display,
                    data=grp['data'],
                    citation=citation,
                )
                counters['concepts'] += 1

            # ── Sheet 3: Sites ───────────────────────
            ws = wb['3_Sites']
            hmap = header_map(ws)
            period_idx = find_col(hmap, 'Period/Phase', 'Culture/Phase')
            for d in rows_as_dicts(ws):
                topic_obj = get_topic(d.get('topic', ''))
                sub_topic_obj = get_subtopic(topic_obj, d.get('microtopic', ''))
                # period column may be 'period/phase' or 'culture/phase'
                period_val = ''
                for cand in ['period/phase', 'culture/phase']:
                    if cand in d:
                        period_val = d[cand]
                        break
                Site.objects.create(
                    chapter=chapter,
                    name=d.get('site', ''),
                    state_region=d.get('state/region', ''),
                    period=period_val,
                    topic=topic_obj,
                    sub_topic=sub_topic_obj,
                    key_findings=d.get('key finding', ''),
                    citation=d.get('citation', ''),
                )
                counters['sites'] += 1

            # ── Sheet 4: KeyFacts ────────────────────
            ws = wb['4_KeyFacts']
            for order, d in enumerate(rows_as_dicts(ws)):
                topic_obj = get_topic(d.get('topic', ''))
                sub_topic_obj = get_subtopic(topic_obj, d.get('microtopic', ''))
                Fact.objects.create(
                    chapter=chapter,
                    topic=topic_obj,
                    sub_topic=sub_topic_obj,
                    text=d.get('fact', ''),
                    citation=d.get('citation', ''),
                    source_sheet='KeyFacts',
                    sort_order=order,
                )
                counters['keyfacts'] += 1

            # ── Sheet 5: State_Specific ──────────────
            ws = wb['5_State_Specific']
            for order, d in enumerate(rows_as_dicts(ws)):
                topic_obj = get_topic(d.get('topic', ''))
                sub_topic_obj = get_subtopic(topic_obj, d.get('microtopic', ''))
                fact = Fact.objects.create(
                    chapter=chapter,
                    topic=topic_obj,
                    sub_topic=sub_topic_obj,
                    text=d.get('fact', ''),
                    citation=d.get('citation', ''),
                    source_sheet='State_Specific',
                    sort_order=order,
                )
                # Link to State if we can match the state name
                state_name = d.get('state', '')
                if state_name:
                    states = State.objects.filter(name__icontains=state_name)
                    if states.exists():
                        fact.state_relevance.add(*states)
                counters['state_specific'] += 1

            # ── Sheet 6: Society ─────────────────────
            ws = wb['6_Society']
            hmap = header_map(ws)
            for order, d in enumerate(rows_as_dicts(ws)):
                topic_obj = get_topic(d.get('topic', ''))
                sub_topic_obj = get_subtopic(topic_obj, d.get('microtopic', ''))
                # Build fact text: "Category: Detail"
                category = d.get('category', '')
                detail = d.get('detail', '')
                text = f"[{category}] {detail}" if category else detail
                # Extra context from period or region/site column
                extra = ''
                for cand in ['period', 'region/site']:
                    if cand in d and d[cand]:
                        extra = d[cand]
                        break
                if extra:
                    text = f"{text} ({extra})"
                Fact.objects.create(
                    chapter=chapter,
                    topic=topic_obj,
                    sub_topic=sub_topic_obj,
                    text=text,
                    citation=d.get('citation', ''),
                    source_sheet='Society',
                    sort_order=order,
                )
                counters['society'] += 1

            # ── Sheet 7: Images ──────────────────────
            ws = wb['7_Images']
            hmap = header_map(ws)
            for d in rows_as_dicts(ws):
                topic_obj = get_topic(d.get('topic', ''))
                ref_code = d.get('image ref', '')
                if not ref_code:
                    continue
                # Make ref_code unique per chapter
                full_ref = f"{chapter_slug}-{ref_code}"
                Visual.objects.create(
                    chapter=chapter,
                    ref_code=full_ref,
                    description=d.get('description', ''),
                    source_book=d.get('source', ''),
                    topic=topic_obj,
                )
                counters['images'] += 1

            # ── Sheet 8: Terms ───────────────────────
            ws = wb['8_Terms']
            for d in rows_as_dicts(ws):
                topic_obj = get_topic(d.get('topic', ''))
                term_name = d.get('term', '')
                if not term_name:
                    continue
                GlossaryTerm.objects.create(
                    chapter=chapter,
                    term=term_name,
                    definition=d.get('definition', ''),
                    citation=d.get('citation', ''),
                    topic=topic_obj,
                )
                counters['terms'] += 1

            # ── Sheet 9: ExamAnalysis ────────────────
            ws = wb['9_ExamAnalysis']
            for d in rows_as_dicts(ws):
                topic_obj = get_topic(d.get('topic', ''))
                category = d.get('category', '')
                ExamIntelEntry.objects.create(
                    chapter=chapter,
                    category=category,
                    topic=topic_obj,
                    detail=d.get('detail', ''),
                    citation=d.get('citation', ''),
                )
                counters['exam_analysis'] += 1

            # ── Sheet 10: Exercises ──────────────────
            ws = wb['10_Exercises']
            for d in rows_as_dicts(ws):
                topic_obj = get_topic(d.get('topic', ''))
                Exercise.objects.create(
                    chapter=chapter,
                    exercise_type=d.get('type', ''),
                    topic=topic_obj,
                    question=d.get('question', ''),
                    source=d.get('source', ''),
                )
                counters['exercises'] += 1

        wb.close()

        # ── Summary ──────────────────────────────
        self.stdout.write(self.style.SUCCESS(f"\n[OK] Chapter '{chapter_name}' uploaded."))
        self.stdout.write(f"    Topics created:     {counters['topics_created']}")
        self.stdout.write(f"    SubTopics created:  {counters['subtopics_created']}")
        self.stdout.write(f"    Timeline events:    {counters['timeline']}")
        self.stdout.write(f"    Concept matrices:   {counters['concepts']}")
        self.stdout.write(f"    Sites:              {counters['sites']}")
        self.stdout.write(f"    KeyFacts:           {counters['keyfacts']}")
        self.stdout.write(f"    State-specific:     {counters['state_specific']}")
        self.stdout.write(f"    Society facts:      {counters['society']}")
        self.stdout.write(f"    Visuals:            {counters['images']}")
        self.stdout.write(f"    Glossary terms:     {counters['terms']}")
        self.stdout.write(f"    Exam intel entries:  {counters['exam_analysis']}")
        self.stdout.write(f"    Exercises:          {counters['exercises']}")
        total = sum(v for k, v in counters.items()
                    if k not in ('topics_created', 'subtopics_created'))
        self.stdout.write(self.style.SUCCESS(f"    TOTAL records:      {total}"))

    def _dry_run(self, wb, sheets, chapter_name):
        """Preview row counts per sheet."""
        self.stdout.write(f"\nChapter: {chapter_name}")
        total = 0
        for sn in sheets:
            ws = wb[sn]
            count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True)
                        if any(c is not None for c in _))
            total += count
            hmap = header_map(ws)
            self.stdout.write(f"  {sn:20s}  {count:4d} rows  headers={list(hmap.keys())}")
        self.stdout.write(self.style.SUCCESS(f"  {'TOTAL':20s}  {total:4d} rows"))
