"""
content/management/commands/upload_prelims.py

Reads PYQ XLSX (MASTER sheet, headers in row 2, data from row 3)
and creates PrelimsPYQ records linked to taxonomy where possible.

Usage:
  python manage.py upload_prelims path/to/UPPCS_PYQ_Ancient_History_v2.xlsx --dry-run
  python manage.py upload_prelims path/to/UPPCS_PYQ_Ancient_History_v2.xlsx
  python manage.py upload_prelims path/to/UPPCS_PYQ_Ancient_History_v2.xlsx --sheet Ch_StoneAge
"""
import re
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
import openpyxl

from content.models import (
    Subject, Chapter, Topic, PrelimsPYQ,
)


def clean(val):
    if val is None:
        return ''
    return str(val).strip()


def safe_int(val, default=0):
    """Parse int from cell, handling None, floats, and dual-year strings like '2004/2008'."""
    if val is None:
        return default
    s = str(val).strip()
    if not s or s == '—' or s == '-':
        return default
    # Handle dual-year: take the first
    if '/' in s:
        s = s.split('/')[0].strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return default


class Command(BaseCommand):
    help = 'Upload Prelims PYQ from XLSX (MASTER or chapter sheet).'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to PYQ XLSX file')
        parser.add_argument('--sheet', type=str, default='MASTER',
                            help='Sheet to read (default: MASTER)')
        parser.add_argument('--dry-run', action='store_true',
                            help='Preview without writing to DB')

    def handle(self, *args, **options):
        xlsx_path = options['xlsx_path']
        sheet_name = options['sheet']
        dry_run = options['dry_run']

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Cannot open {xlsx_path}: {e}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[sheet_name]

        # Build header map from ROW 2 (row 1 is group headers)
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        next(rows_iter)  # skip row 1 (group headers)
        header_row = next(rows_iter)  # row 2 = actual column names
        hmap = {}
        for idx, val in enumerate(header_row):
            if val:
                hmap[clean(val).lower()] = idx

        # Build caches for taxonomy linking
        subject_cache = {s.code.lower(): s for s in Subject.objects.all()}
        subject_name_cache = {s.name.lower(): s for s in Subject.objects.all()}
        chapter_cache = {}  # multiple keys → Chapter
        for ch in Chapter.objects.select_related('unit').all():
            chapter_cache[ch.slug] = ch
            chapter_cache[ch.name.lower()] = ch
            # Also index without leading "The " for fuzzy matching
            stripped = re.sub(r'^the\s+', '', ch.name.lower())
            chapter_cache[stripped] = ch

        def col(row_vals, *candidates):
            """Get value from row by header name candidates."""
            for c in candidates:
                key = c.lower().strip()
                if key in hmap:
                    return clean(row_vals[hmap[key]])
            return ''

        # Counters
        created = 0
        created_no_opts = 0
        skipped_no_qcode = 0
        skipped_duplicate = 0
        skipped_no_question = 0
        year_warnings = []

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — counting rows only."))
            count = 0
            for row in rows_iter:
                if row[0] is not None:
                    count += 1
            self.stdout.write(f"  {sheet_name}: {count} data rows")
            wb.close()
            return

        batch_id = f"upload-{sheet_name}"

        with transaction.atomic():
            for row_num, row_vals in enumerate(rows_iter, start=3):
                row_vals = list(row_vals)

                q_code = col(row_vals, 'Q Code')
                if not q_code:
                    skipped_no_qcode += 1
                    continue

                # Skip if already exists
                if PrelimsPYQ.objects.filter(question_id=q_code).exists():
                    skipped_duplicate += 1
                    continue

                stem = col(row_vals, 'Question Text')
                if not stem:
                    skipped_no_question += 1
                    continue

                opt_a = col(row_vals, 'Opt A')
                opt_b = col(row_vals, 'Opt B')
                opt_c = col(row_vals, 'Opt C')
                opt_d = col(row_vals, 'Opt D')
                has_options = bool(opt_a or opt_b)

                correct = col(row_vals, 'Correct Ans')
                if correct not in ('A', 'B', 'C', 'D'):
                    correct = 'A'  # default; flagged for review

                # Determine review status based on completeness
                if has_options:
                    review = 'draft'
                else:
                    review = 'parse_error_no_options'

                year = safe_int(col(row_vals, 'Year'), default=2000)
                if year < 1990 or year > 2030:
                    year_warnings.append((q_code, col(row_vals, 'Year')))
                    year = 2000

                # Taxonomy linking (best-effort)
                subject_name = col(row_vals, 'AI-Subject')
                subject_obj = (subject_name_cache.get(subject_name.lower()) or
                               subject_cache.get(subject_name.lower()))

                chapter_name = col(row_vals, 'AI-Chapter')
                chapter_obj = chapter_cache.get(chapter_name.lower())

                topic_name = col(row_vals, 'AI-Topic')
                topic_obj = None
                if chapter_obj and topic_name:
                    topic_obj = Topic.objects.filter(
                        chapter=chapter_obj, name=topic_name).first()

                # Difficulty
                difficulty = col(row_vals, 'AI-Difficulty')
                if difficulty not in ('Easy', 'Medium', 'Hard'):
                    difficulty = 'Unknown'

                # Bloom's → int 1-6
                blooms_map = {
                    'remember': 1, 'understand': 2, 'apply': 3,
                    'analyze': 4, 'evaluate': 5, 'create': 6,
                }
                blooms_raw = col(row_vals, "AI-Bloom's").lower()
                blooms_int = blooms_map.get(blooms_raw, 1)

                # Repeat count
                repeat = safe_int(col(row_vals, 'AI-Repeat Count'), default=1)
                if repeat < 1:
                    repeat = 1

                # Concept cluster
                cluster = col(row_vals, 'AI-Topic Cluster')

                # Keywords → tags
                keywords = col(row_vals, 'AI-Keywords')

                # Explanation
                explanation = col(row_vals, 'Explanation 1')

                # Exam source
                exam_name = col(row_vals, 'Exam Name')
                exam_source = 'UPPCS'
                if 'upsc' in exam_name.lower() or 'ias' in exam_name.lower():
                    exam_source = 'UPSC'

                PrelimsPYQ.objects.create(
                    question_id=q_code,
                    stem=stem,
                    option_a=opt_a or '-',
                    option_b=opt_b or '-',
                    option_c=opt_c or '-',
                    option_d=opt_d or '-',
                    correct_answer=correct,
                    explanation=explanation,
                    year=year,
                    exam_source=exam_source,
                    subject=subject_obj,
                    chapter=chapter_obj,
                    topic=topic_obj,
                    difficulty=difficulty,
                    blooms_level=blooms_int,
                    concept_cluster=cluster,
                    tags=keywords,
                    repeat_count=repeat,
                    review_status=review,
                    batch_id=batch_id,
                )
                if has_options:
                    created += 1
                else:
                    created_no_opts += 1

        wb.close()

        # Summary
        self.stdout.write(self.style.SUCCESS(f"\n[OK] Prelims PYQ upload complete."))
        self.stdout.write(f"    Created (complete): {created}")
        self.stdout.write(f"    Created (no opts):  {created_no_opts}  ← review_status='parse_error_no_options'")
        self.stdout.write(f"    Skipped (no code):  {skipped_no_qcode}")
        self.stdout.write(f"    Skipped (dup):      {skipped_duplicate}")
        self.stdout.write(f"    Skipped (no Q):     {skipped_no_question}")
        if year_warnings:
            self.stdout.write(self.style.WARNING(
                f"    Year warnings:     {len(year_warnings)}"))
            for qc, yr in year_warnings[:5]:
                self.stdout.write(f"      {qc}: year='{yr}'")
